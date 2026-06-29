import openpyxl
wb = openpyxl.load_workbook('results/results.xlsx')
ws = wb.active
row = ws.max_row + 1
d = {}
d[1] = 'Ayop et al. (2018) - Components sizing of photovoltaic stand-alone system based on loss of power supply probability'
d[2] = 'Paper uses battery storage (57 banks baseline, 62 RIM) instead of hydrogen storage. Battery SOC limits: 20-80%. Battery charging efficiency: 80%. Battery life: 7 years. PV array: 5x5 (25 modules of 315W). Rated power per PV array stated as "7875 kW" in text -- likely typo for 7875 W (7.875 kW). Total annual load (296,745 kWh) and daily load have discrepancy (813 kWh p.7 vs 813,000 kWh Table 5). Baseline (LCC): 21 PV arrays, 57 battery banks, LPSP=0.812%. RIM: 22 PV arrays, 62 battery banks, LPSP=0.18%. Interest rate/discount factor referenced but not given. GCM initial cost: RM 1,264,997. Annual maintenance (LCC): RM 247,774. No LCOE, COE reported. No WT, H2, FC, EL, DG.'
d[3] = '296,745 kW h'
d[5] = '0.812%'
d[7] = '315 W'
d[23] = '90%'
d[25] = 'RM 33,900'
d[31] = 'RM 125,621'
d[60] = '25 years'
d[66] = '13 years'
d[67] = 21
for c, v in d.items():
    ws.cell(row=row, column=c, value=v)
wb.save('results/results.xlsx')
print(f'Appended at row {row}')
