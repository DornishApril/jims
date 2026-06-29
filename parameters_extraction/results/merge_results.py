import json
import os
import glob

try:
    from openpyxl import Workbook
except ImportError:
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
    exit(1)

# Define the exact column order from prompt.txt
COLUMNS = [
    'Paper Name', 'NOTES', 'TOTAL ANNUAL LOAD DEMAND', 'LCOE', 'LPSP', 'COE',
    'rated_PV', 'v_cut_in', 'v_rated', 'rated_power', 'Cap_H2', 'Cap_FC', 'Cap_EL', 'Cap_DG',
    'H_min_percentage', 'H_max_percentage', 'f_0', 'f_1',
    'eta_PV', 'eta_FC', 'eta_EL', 'eta_INVT', 'H2_LHV',
    'c_PV', 'c_WT', 'c_H2', 'c_FC_cap', 'c_EL_cap', 'c_DG_cap', 'c_INVT',
    'c_FC', 'c_DG', 'c_EL', 'c_DG_FUEL',
    'om_PV', 'om_WT', 'om_H2', 'om_FC', 'om_EL', 'om_DG', 'om_INVT',
    'rc_PV', 'rc_WT', 'rc_H2', 'rc_FC', 'rc_EL', 'rc_DG', 'rc_INVT',
    'e_FC', 'e_DG', 'e_EL',
    'T_life', 'r', 'p_grid',
    'A_PV', 'P_DG_min',
    'life_PV', 'life_WT', 'life_H2', 'life_FC', 'life_EL', 'life_DG', 'life_INVT',
    'N_PV', 'N_WT', 'N_H2', 'N_FC', 'N_EL', 'N_DG'
]

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PER_PAPER_DIR = os.path.join(SCRIPT_DIR, 'per_paper')
OUTPUT_FILE = os.path.join(SCRIPT_DIR, 'results.xlsx')

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = 'Parameter Extraction'

# Write header
for col_idx, col_name in enumerate(COLUMNS, 1):
    ws.cell(row=1, column=col_idx, value=col_name)

# Read all per-paper JSON files
json_files = sorted(glob.glob(os.path.join(PER_PAPER_DIR, '*.json')))
print(f"[INFO] Found {len(json_files)} per-paper JSON files")

row = 2
for jf in json_files:
    try:
        with open(jf, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Handle both single dict and list of dicts
        if isinstance(data, list):
            entries = data
        else:
            entries = [data]
        
        for entry in entries:
            # Paper name from entry or filename
            paper_name = entry.get('Paper Name', '') or entry.get('paper_name', '') or os.path.basename(jf).replace('.json', '')
            ws.cell(row=row, column=1, value=paper_name)
            
            for col_idx, col_name in enumerate(COLUMNS[1:], 2):
                val = entry.get(col_name, None)
                if val is not None and val != '' and val != 'N/A':
                    ws.cell(row=row, column=col_idx, value=val)
            
            print(f"  [{row-1}] {paper_name[:60]}")
            row += 1
            
    except Exception as e:
        print(f"  [ERROR] Failed to read {os.path.basename(jf)}: {e}")

# Save
wb.save(OUTPUT_FILE)
print(f"\n[DONE] Merged {row-2} papers into {OUTPUT_FILE}")
