import openpyxl
wb = openpyxl.load_workbook('results/results.xlsx')
ws = wb.active

# Fix row 8 (Ayop paper) with correct column indices
d = {
    1: 'Ayop et al. (2018) - Components sizing of photovoltaic stand-alone system based on loss of power supply probability',
    2: 'Paper uses battery storage (57 banks baseline, 62 RIM) instead of hydrogen storage. Battery SOC limits: 20-80%. Battery charging efficiency: 80%. Battery life: 7 years. PV array: 5x5 (25 modules of 315W). Rated power per PV array stated as "7875 kW" in text -- likely typo for 7875 W (7.875 kW). Total annual load (296,745 kWh) and daily load have discrepancy (813 kWh p.7 vs 813,000 kWh Table 5). Baseline (LCC): 21 PV arrays, 57 battery banks, LPSP=0.812%. RIM: 22 PV arrays, 62 battery banks, LPSP=0.18%. Interest rate/discount factor referenced but not given. GCM initial cost: RM 1,264,997. Annual maintenance (LCC): RM 247,774. No LCOE, COE reported. No WT, H2, FC, EL, DG.',
    3: '296,745 kW h',
    5: '0.812%',
    7: '315 W',
    22: '90%',
    24: 'RM 33,900',
    30: 'RM 125,621',
    57: '25 years',
    63: '13 years',
    64: 21
}

# Clear existing row 8 first then write
row = 8
for c in range(1, 70):
    ws.cell(row=row, column=c).value = None

for c, v in d.items():
    ws.cell(row=row, column=c, value=v)

wb.save('results/results.xlsx')

# Verify
wb2 = openpyxl.load_workbook('results/results.xlsx')
ws2 = wb2.active
h = [ws2.cell(1,c).value for c in range(1,70)]
d2 = [ws2.cell(8,c).value for c in range(1,70)]
print('Corrected row 8:')
for i in range(69):
    if d2[i] is not None:
        print(f'  col {i+1:2d} {h[i]:40s} = {d2[i]}')
